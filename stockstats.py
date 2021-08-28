
from pprint import pprint
from pathlib import Path
import os
import pandas as pan
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse, sys

from datetime import datetime, timedelta
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import json
import requests
import time
from bs4 import BeautifulSoup
import finviz
import numpy as np
import tweepy
import snscrape.modules.twitter as sntwitter

from iextoken import token







def createSheets(name, book):
    print("create sheets")
    headings = ["Date", "Symbol", "Last", "High Since date", "Date of High", "Last High CD"] # all values must be strings
    ws1 = book.create_sheet(name)
    i = 1
    for item in headings:
        #print(item)
        ws1.cell(row = 1, column = i).value = item
        i=i+1



def getNumDaysStr(date, append, offset):
    lastDate = date
    lastTradeDateUrl = "https://cloud.iexapis.com/stable/ref-data/us/dates/trade/last/1?token="+token
    lastTradeRes = requests.get(lastTradeDateUrl)
    #print(type(lastTradeRes))
    lastTradeDate = lastTradeRes.json()[0]["date"]
    # print(type(lastTradeDate))
    # print(lastTradeDate)


    lastTradeDateDate = datetime.strptime(lastTradeDate, "%Y-%m-%d")

    numdays = (lastTradeDateDate - lastDate).days
    numdays += append + offset
    print("Numdays", numdays)


    rangeToUse = ""

    if (numdays < 32):
        rangeToUse = "1m"
    elif (numdays >= 32 and numdays < 64):
        rangeToUse = "3m"
    elif (numdays >= 64 and numdays < 127):
        rangeToUse = "6m"
    elif (numdays >= 127 and numdays < 254):
        rangeToUse = "1y"
    elif (numdays >= 254 and numdays < 508):
        rangeToUse = "2y"
    elif (numdays >= 508 and numdays < 1270):
        rangeToUse = "5y"
    else:
        rangeToUse = "max"

    return rangeToUse



def getNumDaysStrMax(date, append, offset):
    try:
        lastDate = date


        lastTradeDateUrl = "https://cloud.iexapis.com/stable/ref-data/us/dates/trade/last/1?token="+token
        lastTradeRes = requests.get(lastTradeDateUrl)
        #print(type(lastTradeRes))
        lastTradeDate = lastTradeRes.json()[0]["date"]
        # print(type(lastTradeDate))
        # print(lastTradeDate)


        lastTradeDateDate = datetime.strptime(lastTradeDate, "%Y-%m-%d")

        numdays = (lastTradeDateDate - lastDate).days
        numdays += append + offset
        print("Numdays", numdays)


        rangeToUse = ""

        if (numdays < 30):
            rangeToUse = "1m"
        elif (numdays >= 30 and numdays < 90):
            rangeToUse = "3m"
        elif (numdays >= 90 and numdays < 180):
            rangeToUse = "6m"
        elif (numdays >= 180 and numdays < 365):
            rangeToUse = "1y"
        elif (numdays >= 365 and numdays < 730):
            rangeToUse = "2y"
        elif (numdays >= 730 and numdays < 1825):
            rangeToUse = "5y"
        else:
            rangeToUse = "max"

        return rangeToUse, numdays, lastTradeDateDate
    except Exception as ex:
        print(ex)
        return None


def calcIndicator(url, date, offset):
    try:
        data = requests.get(url).json()
        indicatorArr = data['indicator']
        chart = data['chart']
        index = 0
        for close in chart:
            closeDate = close['date']
            #closeDateConv = closeDate.split("-")[0]+closeDate.split("-")[1]+closeDate.split("-")[2]
            closeDateConvTime = datetime.strptime(closeDate, "%Y-%m-%d")
            #print("index:", index, "val: ", indicatorArr[0][index])
            if (closeDateConvTime == date):
                print(date)
                index-=offset
                val = indicatorArr[0][index]
                #print("Length of indicator: ", len(indicatorArr), "Length of chart: ", len(chart), "index: ", index)
                # if (normalized and val!=None):
                #     val /= fclose
                print("indicator found at ", date, "val= ", val)
                if (val == None):
                    return "#N/A"
                return val
            index+=1
    except Exception as ex:
        print(ex)
        return "#N/A" 


def calcAvgVol(data, date, period):
    try:
        index = 0
        canLoop = False
        volArray = []
        count = 0
        exactVolume = 0
        for close in data:
            if (count >= period):
                break
            closeDate = close['date']
            volume = close['volume']
            if (canLoop):
                volArray.append(float(volume))
                count+=1
            #closeDateConv = closeDate.split("-")[0]+closeDate.split("-")[1]+closeDate.split("-")[2]
            closeDateConvTime = datetime.strptime(closeDate, "%Y-%m-%d")
            #print("index:", index, "val: ", indicatorArr[0][index])
            if (closeDateConvTime == date):
                print("exactdate volcalc", date)
                volArray.append(float(volume))
                exactVolume = float(volume)
                count+=1
                canLoop = True

            index+=1
        volAvg = np.average(volArray)
        volStdDev = np.std(volArray)
        print("Vol avg calculated for period ", count, "avg = ", volAvg, "stddev = ", volStdDev)
        return volAvg, volStdDev, exactVolume
    except Exception as ex:
        print(ex)
        return "#N/A" 



def calcAtr(url, date, atrPeriod, avgAtrPeriod):
    try:
        canLoop = False
        atrArray = []
        count = 0
        exactAtr = 0

        data = requests.get(url).json()
        indicatorArr = data['indicator'][0]
        chart = data['chart']
        indicatorArr.reverse()
        chart.reverse()


        index = 0

        for close in chart:
            if (count >= avgAtrPeriod):
                break
            closeDate = close['date']
            if (canLoop):
                val = float(indicatorArr[index])
                atrArray.append(val)
                count+=1
            #closeDateConv = closeDate.split("-")[0]+closeDate.split("-")[1]+closeDate.split("-")[2]
            closeDateConvTime = datetime.strptime(closeDate, "%Y-%m-%d")
            #print("index:", index, "val: ", indicatorArr[0][index])
            if (closeDateConvTime == date):
                print("exactdate atrcalc", date)
                val = float(indicatorArr[index])
                atrArray.append(val)
                exactAtr = val
                count+=1
                canLoop = True

            index+=1
        atrAvg = np.average(atrArray)
        print("Atr avg calculated for period ", count, "avg = ", atrAvg, "atr = ", exactAtr)
        return atrAvg, exactAtr
    except Exception as ex:
        print(ex)
        return "#N/A" 


def getStockData(sht, calcShOut, calcFloat, calcCountry, calcSector, calcIndustry, minrow, maxrow, calcOverwrite):
    # auth = tweepy.AppAuthHandler(nhLX0Abge9AIkSrBhCXukZMvL, CRopEqY1RiFbiuH3IekyopYKDtIObaRwONIDEJLq6ehilirPb6)
    # api = tweepy.API(auth)
    themesxl = 'stockstats.xlsx'
    xl = load_workbook(filename = themesxl)
    stocks = xl[sht]
    minRow = minrow
    maxRow = maxrow
    rowIndex = minRow
    symbolCol = column_index_from_string("D")
    priceAddedCol = column_index_from_string("F")
    dateAddedCol = column_index_from_string("G")
    lastPriceCol = column_index_from_string("H")
    highestPriceCol = column_index_from_string("I")
    highDateCol = column_index_from_string("J")
    lastHighComputeCol = column_index_from_string("M")
    manualOverrideCol = column_index_from_string("N")
    daysElapsedCol = column_index_from_string("Q")
    shOutCol = column_index_from_string("O")
    floatCol = column_index_from_string("P")
    countryCol = column_index_from_string("R")
    sectorCol = column_index_from_string("S")
    industryCol = column_index_from_string("T")
    socialCol = column_index_from_string("U")
    dateAddedVolumeCol = column_index_from_string("V")


    rowIndex = minRow
    for row in stocks.iter_rows(min_row=minRow, max_row=maxRow):
        symbol = stocks.cell(rowIndex, symbolCol).value
        print(symbol)
        if (symbol != None and stocks.cell(rowIndex, manualOverrideCol).value == None):
            # do last price calc
            lastPrice = stocks.cell(rowIndex, lastPriceCol).value
            url = 'https://cloud.iexapis.com/stable/stock/'+symbol+'/price?token=pk_5e8779cbd4a34bc89a43575e146be124'
            data = requests.get(url).json()
            stocks.cell(rowIndex, lastPriceCol).value = data
            stocks.cell(rowIndex, lastPriceCol).number_format = "$0.00##"


            # do high since date
            stockHigh = stocks.cell(rowIndex, highestPriceCol).value
            if (stockHigh == None):
                stockHigh = stocks.cell(rowIndex, priceAddedCol).value
                stocks.cell(rowIndex, highestPriceCol).value = stocks.cell(rowIndex, priceAddedCol).value
                stocks.cell(rowIndex, highDateCol).value = stocks.cell(rowIndex, dateAddedCol).value
                stocks.cell(rowIndex, highDateCol).number_format = "m/d/yyyy"

            lastHighCompute = stocks.cell(rowIndex, lastHighComputeCol).value
            # now get the date to look back
            date = lastHighCompute
            if (lastHighCompute == None):
                date = stocks.cell(rowIndex, dateAddedCol).value

            try:
                rangetouse, numdays, lastTradeDate = getNumDaysStrMax(date, 0, 0)
            except Exception as ex:
                print(ex)

            if (numdays > 0):
                try:
                    url = 'https://cloud.iexapis.com/beta/stock/market/batch?symbols='+str(symbol)+'&types=chart&chartByDay=true&range=' + str(rangetouse) + '&chartLast='+str(numdays)+'&filter=date,high&token=pk_5e8779cbd4a34bc89a43575e146be124'
                    data = requests.get(url).json()
                    currHigh = stockHigh
                    highDate = ""
                    minDate = stocks.cell(rowIndex, dateAddedCol).value
                    #print(url)

                    for stock in data.items():
                        for chart in stock[1].items():
                            for close in chart[1]:
                                highDateConv = datetime.strptime(close['date'], "%Y-%m-%d")
                                if ((highDateConv - minDate).days <= 0):
                                    break
                                currHigh = float(close['high'])
                                if (currHigh > stockHigh):
                                    stockHigh = currHigh
                                    highDate = close['date']



                    if (stocks.cell(rowIndex, dateAddedVolumeCol).value == None):
                        url = "https://cloud.iexapis.com/stable/stock/"+str(symbol)+"/chart/date/"+datetime.strftime(minDate, "%Y%m%d")+"?chartByDay=true&chartCloseOnly=true&token="+token

                        data1 = requests.get(url).json()

                        calc = float(data1[0]["volume"])
                        stocks.cell(rowIndex, dateAddedVolumeCol).value = calc

                    if (highDate != ""):
                        print("new high found ", stockHigh, highDate)
                        stocks.cell(rowIndex, highestPriceCol).value = stockHigh
                        stocks.cell(rowIndex, highestPriceCol).number_format = "$0.00##"
                        highDateConv = datetime.strptime(highDate, "%Y-%m-%d")
                        elapsed = (highDateConv - stocks.cell(rowIndex, dateAddedCol).value).days
                        stocks.cell(rowIndex, highDateCol).value = highDateConv
                        stocks.cell(rowIndex, highDateCol).number_format = "m/d/yyyy"
                        stocks.cell(rowIndex, daysElapsedCol).value = elapsed
                    else:
                        print("no new high for this period")
                    stocks.cell(rowIndex, lastHighComputeCol).value = lastTradeDate
                    stocks.cell(rowIndex, lastHighComputeCol).number_format = "m/d/yyyy"
                except Exception as ex:
                    print(ex)

            else:
                print("No data to load")


            # do key stats
            try:
                if (stocks.cell(rowIndex, socialCol).value == None):
                    #https://github.com/igorbrigadir/twitter-advanced-search
                    #tweetarchive
                    minDate = stocks.cell(rowIndex, dateAddedCol).value
                    dBack = minDate - timedelta(days=5)
                    # tweets=tweepy.Cursor(api.search_full_archive,environment_name="tweetarchive", query="$"+symbol, fromDate=date_since_pro).items(numTweets)
                    # api.search_full_archive("tweetarchive", query="$"+symbol, fromDate=date_since_pro).items(numTweets)

                    minDateStr = datetime.strftime(minDate, "%Y-%m-%d")
                    firstDateStr = datetime.strftime(dBack, "%Y-%m-%d")
                    #$aht since:2021-01-01 until:2021-05-01
                    # Creating list to append tweet data to
                    tweets_list2 = []
                    searchStr = "$"+symbol+" since:"+firstDateStr+ " until:"+minDateStr
                    print(searchStr)
                    # Using TwitterSearchScraper to scrape data and append tweets to list
                    for i,tweet in enumerate(sntwitter.TwitterSearchScraper(searchStr).get_items()):
                        if i>500:
                            break
                        tweets_list2.append([tweet.date, tweet.id, tweet.content, tweet.user.username])
                        
                    # Creating a dataframe from the tweets list above
                    tweets_df2 = pan.DataFrame(tweets_list2, columns=['Datetime', 'Tweet Id', 'Text', 'Username'])
                    print("tweets: ",tweets_df2.size)
                    stocks.cell(rowIndex, socialCol).value = tweets_df2.size



                if (calcShOut and (calcOverwrite or stocks.cell(rowIndex, shOutCol).value == None)):
                    # shares out
                    url = "https://cloud.iexapis.com/stable/stock/"+symbol+"/stats/sharesOutstanding?token="+token
                    data = requests.get(url).json()
                    stocks.cell(rowIndex, shOutCol).value = data
                    print(data)



                if (calcSector and (calcOverwrite or stocks.cell(rowIndex, sectorCol).value == None)):
                    # sector
                    url = "https://cloud.iexapis.com/stable/stock/"+symbol+"/company?filter=sector&token="+token
                    data = requests.get(url).json()
                    stocks.cell(rowIndex, sectorCol).value = data['sector']
                    print(data['sector'])
                    #stocks.cell(rowIndex, floatCol).value = floatVar
                    
                fin = finviz.get_stock(symbol)
                if (calcFloat and (calcOverwrite or stocks.cell(rowIndex, floatCol).value == None)):
                    # shares out
                    floatVar = fin["Shs Float"]
                    if (floatVar.endswith("M")):
                        floatVar = float(floatVar.split("M")[0]) * 1000000
                    elif (floatVar.endswith("B")):
                        floatVar = float(floatVar.split("B")[0]) * 1000000000
                    elif (floatVar.endswith("K")):
                        floatVar = float(floatVar.split("K")[0]) * 1000

                    if (floatVar != "-"):
                        stocks.cell(rowIndex, floatCol).value = floatVar
                        print(floatVar)

                if (calcCountry and (calcOverwrite or stocks.cell(rowIndex, countryCol).value == None)):
                    # shares out
                    country = fin['Country']
                    print(country)
                    stocks.cell(rowIndex, countryCol).value = country


                if (calcIndustry and (calcOverwrite or stocks.cell(rowIndex, industryCol).value == None)):
                    # industry
                    industry = fin['Industry']
                    print(industry)
                    stocks.cell(rowIndex, industryCol).value = industry
            except Exception as ex:
                print(ex)



            









        rowIndex+=1
        #xl.save(themesxl)
    try:
        xl.save(themesxl)
    except Exception as ex:
        failed = True
        while(failed):
            print(ex)
            stri = input("Save failed, try again...? Press any key to try again")
            try:
                xl.save(themesxl)
                failed = False
            except Exception as ex:
                failed = True
    xl.close()






def loadIndicators(sht, minrow, maxrow):

    themesxl = 'stockstats.xlsx'


    # bs for future use


    xl = load_workbook(filename = themesxl)
    sheet_names = xl.sheetnames  # a list of existing sheet names

    stocks = xl[sht]

    themeRow = 1
    minRow = minrow
    maxRow = maxrow



    minCol = column_index_from_string('Y')
    maxCol = column_index_from_string("CQ")
    symbolCol = column_index_from_string("D")
    dateAddedCol = column_index_from_string("G")


    fClose = column_index_from_string("X")

    volumeatclosecol = column_index_from_string("V")

    #sma
    sma200Col = column_index_from_string("Y")
    sma100Col = column_index_from_string("Z")
    sma50Col = column_index_from_string("AA")
    dma50Col = column_index_from_string("AB")
    sma20Col = column_index_from_string("AC")
    sma14Col = column_index_from_string("AD")
    sma8Col = column_index_from_string("AE")

    #slope
    lrs200Col = column_index_from_string("AF")
    lrs100Col = column_index_from_string("AG")
    lrs50Col = column_index_from_string("AH")
    lrs50dCol = column_index_from_string("AI")
    lrs20Col = column_index_from_string("AJ")
    lrs14Col = column_index_from_string("AK")
    lrs8Col = column_index_from_string("AL")


    #diff
    diff200_100 = column_index_from_string("AM")
    diff200_50 = column_index_from_string("AN")
    diff200_50d = column_index_from_string("AO")
    diff200_20 = column_index_from_string("AP")
    diff200_14 = column_index_from_string("AQ")
    diff200_8 = column_index_from_string("AR")

    diff100_50 = column_index_from_string("AS")
    diff100_50d = column_index_from_string("AT")
    diff100_20 = column_index_from_string("AU")
    diff100_14 = column_index_from_string("AV")
    diff100_8 = column_index_from_string("AW")

    diff50d_50 = column_index_from_string("AX")
    diff50d_20 = column_index_from_string("AY")
    diff50d_14 = column_index_from_string("AZ")
    diff50d_8 = column_index_from_string("BA")

    diff50_20 = column_index_from_string("BB")
    diff50_14 = column_index_from_string("BC")
    diff50_8 = column_index_from_string("BD")

    diff20_14 = column_index_from_string("BE")
    diff20_8 = column_index_from_string("BF")

    diff14_8 = column_index_from_string("BG")

    diffall = column_index_from_string("BH")

    volumeAvg5Col = column_index_from_string("BK")
    volumeAvg10Col = column_index_from_string("BL")
    volumeAvg20Col = column_index_from_string("BM")
    volumeAvg50Col = column_index_from_string("BN")
    volumeAvg100Col = column_index_from_string("BO")
    volumeAvg200Col = column_index_from_string("BP")
    rsi14Col = column_index_from_string("BQ")
    unfilledGapAboveCol = column_index_from_string("BR")
    unfilledGapBelowCol = column_index_from_string("BS")
    accuDistLineCol = column_index_from_string("BT")
    accuDistOscCol = column_index_from_string("BU")
    atrCol = column_index_from_string("BV")
    avgAtrCol = column_index_from_string("BW")
    deltaAvgAtr_AtrCol = column_index_from_string("BX")
    chaikinVol10Col = column_index_from_string("BY")
    chaikinVol50Col = column_index_from_string("BZ")
    bbandsUpperCol = column_index_from_string("CA")
    bbandsLowerCol = column_index_from_string("CB")
    bbandsMiddleCol = column_index_from_string("CC")
    bbandsDeltaCol = column_index_from_string("CD")
    macdValueCol = column_index_from_string("CE")
    macdSignalCol = column_index_from_string("CF")
    macdHistoCol = column_index_from_string("CG")
    macdValueSlopeCol = column_index_from_string("CH")
    high52WkCol = column_index_from_string("CI")
    high52WkPercCol = column_index_from_string("CJ")
    low52WkCol = column_index_from_string("CK")
    low52WkPercCol = column_index_from_string("CL")
    onBalanceVolCol = column_index_from_string("CM")
    variance10Col = column_index_from_string("CN")
    variance20Col = column_index_from_string("CO")
    variance50Col = column_index_from_string("CP")
    williamsADCol = column_index_from_string("CQ")




    # loop column in combined sheet
    rowIndex = minRow
    for row in stocks.iter_rows(min_row=minRow, max_row=maxRow):
        colIndex = minCol
        first = 0
        second = 0
        third = 0
        fourth = 0
        fifth = 0
        sixth = 0

        seventh = 0
        eighth = 0
        ninth = 0
        tenth = 0
        eleventh = 0

        twelveth = 0
        thirteenth = 0
        fourtheenth = 0
        fifteenth = 0

        sixteenth = 0
        seventeenth = 0
        eigtheenth = 0

        nineteenth = 0
        twentieth = 0

        twentifirst = 0
        canContinue = True
        for column in stocks.iter_cols(min_row=minRow, min_col=minCol, max_col=maxCol):
        # set theme name to theme value

            symbol = stocks.cell(rowIndex, symbolCol).value
            date = stocks.cell(rowIndex, dateAddedCol).value
            if (stocks.cell(rowIndex, fClose).value == None and canContinue):
                try:
                    url = "https://cloud.iexapis.com/stable/stock/"+symbol+"/chart/date/"+date.strftime("%Y%m%d")+"?chartByDay=true&chartCloseOnly=false&token="+token
                    data = requests.get(url).json()
                    print(url, data)
                    stocks.cell(rowIndex, fClose).value = data[0]['fClose']
                    canContinue = True
                except Exception as ex:
                    print(ex)
                    canContinue = False
            if (stocks.cell(rowIndex, colIndex).value == None and stocks.cell(rowIndex, colIndex).value != "#N/A" and canContinue):
                #print(rowIndex, get_column_letter(colIndex))

                if (colIndex == sma200Col):
                    rangeToUse = getNumDaysStr(date, 200, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/sma?range="+rangeToUse+"&indicatorOnly=false&input1=200&token="+token
                    # '=IF($A${0}="Total%", {1}, IF($A${2}="(+/-)", {3}, {4}))'.format(typeRow, combinedSum, typeRow, combinedSign/count, combinedSign)
                    #print(url)
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = calc



                elif (colIndex == sma100Col):
                    rangeToUse = getNumDaysStr(date, 100,0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/sma?range="+rangeToUse+"&indicatorOnly=false&input1=100&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = calc
                elif (colIndex == sma50Col): 
                    rangeToUse = getNumDaysStr(date, 50, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/sma?range="+rangeToUse+"&indicatorOnly=false&input1=50&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = calc
                elif (colIndex == dma50Col):
                    rangeToUse = getNumDaysStr(date, 60, 10)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/sma?range="+rangeToUse+"&indicatorOnly=false&input1=50&token="+token
                    calc = calcIndicator(url, date, 10)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = calc
                elif (colIndex == sma20Col):
                    rangeToUse = getNumDaysStr(date, 20, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/sma?range="+rangeToUse+"&indicatorOnly=false&input1=20&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = calc
                elif (colIndex == sma14Col):
                    rangeToUse = getNumDaysStr(date, 14, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/sma?range="+rangeToUse+"&indicatorOnly=false&input1=14&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = calc
                elif (colIndex == sma8Col):
                    rangeToUse = getNumDaysStr(date, 8, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/sma?range="+rangeToUse+"&indicatorOnly=false&input1=8&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = calc





                elif (colIndex == lrs200Col):
                    rangeToUse = getNumDaysStr(date, 200, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/linregslope?range="+rangeToUse+"&indicatorOnly=false&input1=200&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(calc, rowIndex, calc)

                elif (colIndex == lrs100Col):
                    rangeToUse = getNumDaysStr(date, 100, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/linregslope?range="+rangeToUse+"&indicatorOnly=false&input1=100&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(calc, rowIndex, calc)
                elif (colIndex == lrs50Col): 
                    rangeToUse = getNumDaysStr(date, 50, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/linregslope?range="+rangeToUse+"&indicatorOnly=false&input1=50&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(calc, rowIndex, calc)
                elif (colIndex == lrs50dCol):
                    rangeToUse = getNumDaysStr(date, 60, 10)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/linregslope?range="+rangeToUse+"&indicatorOnly=false&input1=50&token="+token
                    calc = calcIndicator(url, date, 10)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(calc, rowIndex, calc)
                elif (colIndex == lrs20Col):
                    rangeToUse = getNumDaysStr(date, 20, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/linregslope?range="+rangeToUse+"&indicatorOnly=false&input1=20&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(calc, rowIndex, calc)
                elif (colIndex == lrs14Col):
                    rangeToUse = getNumDaysStr(date, 14, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/linregslope?range="+rangeToUse+"&indicatorOnly=false&input1=14&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(calc, rowIndex, calc)
                elif (colIndex == lrs8Col):
                    rangeToUse = getNumDaysStr(date, 8, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/linregslope?range="+rangeToUse+"&indicatorOnly=false&input1=8&token="+token
                    calc = calcIndicator(url, date, 0)
                    if (calc != None):
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(calc, rowIndex, calc)




                # diff200_100 = column_index_from_string("AM")
                # diff200_50 = column_index_from_string("AN")
                # diff200_50d = column_index_from_string("AO")
                # diff200_20 = column_index_from_string("AP")
                # diff200_14 = column_index_from_string("AQ")
                # diff200_8 = column_index_from_string("AR")

                # diff100_50 = column_index_from_string("AS")
                # diff100_50d = column_index_from_string("AT")
                # diff100_20 = column_index_from_string("AU")
                # diff100_14 = column_index_from_string("AV")
                # diff100_8 = column_index_from_string("AW")

                # diff50d_50 = column_index_from_string("AX")
                # diff50d_20 = column_index_from_string("AY")
                # diff50d_14 = column_index_from_string("AZ")
                # diff100_8 = column_index_from_string("BA")

                # diff50_20 = column_index_from_string("BB")
                # diff50_14 = column_index_from_string("BC")
                # diff50_8 = column_index_from_string("BD")

                # diff20_14 = column_index_from_string("BE")
                # diff20_8 = column_index_from_string("BF")

                # diff14_8 = column_index_from_string("BG")

                # diffall = column_index_from_string("BH")

                elif (colIndex == diff200_100):
                    first = stocks.cell(rowIndex, sma200Col).value
                    second = stocks.cell(rowIndex, sma100Col).value
                    print(first, second)
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            first = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff200_50):
                    first = stocks.cell(rowIndex, sma200Col).value
                    second = stocks.cell(rowIndex, sma50Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            second = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff200_50d):
                    first = stocks.cell(rowIndex, sma200Col).value
                    second = stocks.cell(rowIndex, dma50Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            third = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff200_20):
                    first = stocks.cell(rowIndex, sma200Col).value
                    second = stocks.cell(rowIndex, sma20Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            fourth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff200_14):
                    first = stocks.cell(rowIndex, sma200Col).value
                    second = stocks.cell(rowIndex, sma14Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            fifth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff200_8):
                    first = stocks.cell(rowIndex, sma200Col).value
                    second = stocks.cell(rowIndex, sma8Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            sixth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff100_50):
                    first = stocks.cell(rowIndex, sma100Col).value
                    second = stocks.cell(rowIndex, sma50Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            seventh = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff100_50d):
                    first = stocks.cell(rowIndex, sma100Col).value
                    second = stocks.cell(rowIndex, dma50Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            eighth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff100_20):
                    first = stocks.cell(rowIndex, sma100Col).value
                    second = stocks.cell(rowIndex, sma20Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            ninth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff100_14):
                    first = stocks.cell(rowIndex, sma100Col).value
                    second = stocks.cell(rowIndex, sma14Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            tenth = numerator
                        except Exception as ex:
                            print(ex)


                elif (colIndex == diff100_8):
                    first = stocks.cell(rowIndex, sma100Col).value
                    second = stocks.cell(rowIndex, sma8Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            eleventh = numerator
                        except Exception as ex:
                            print(ex)


                elif (colIndex == diff50d_50):
                    first = stocks.cell(rowIndex, dma50Col).value
                    second = stocks.cell(rowIndex, sma50Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            twelveth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff50d_20):
                    first = stocks.cell(rowIndex, dma50Col).value
                    second = stocks.cell(rowIndex, sma20Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            thirteenth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff50d_14):
                    first = stocks.cell(rowIndex, dma50Col).value
                    second = stocks.cell(rowIndex, sma14Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            fourtheenth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff50d_8):
                    first = stocks.cell(rowIndex, dma50Col).value
                    second = stocks.cell(rowIndex, sma8Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            fifteenth = numerator
                        except Exception as ex:
                            print(ex)



                elif (colIndex == diff50_20):
                    first = stocks.cell(rowIndex, sma50Col).value
                    second = stocks.cell(rowIndex, sma20Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            sixteenth = numerator
                        except Exception as ex:
                            print(ex)


                elif (colIndex == diff50_14):
                    first = stocks.cell(rowIndex, sma50Col).value
                    second = stocks.cell(rowIndex, sma14Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            seventeenth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff50_8):
                    first = stocks.cell(rowIndex, sma50Col).value
                    second = stocks.cell(rowIndex, sma8Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            eigtheenth = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diff20_14):
                    first = stocks.cell(rowIndex, sma20Col).value
                    second = stocks.cell(rowIndex, sma14Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            nineteenth = numerator
                        except Exception as ex:
                            print(ex)


                elif (colIndex == diff20_8):
                    first = stocks.cell(rowIndex, sma20Col).value
                    second = stocks.cell(rowIndex, sma8Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            twentieth = numerator
                        except Exception as ex:
                            print(ex)


                elif (colIndex == diff14_8):
                    first = stocks.cell(rowIndex, sma14Col).value
                    second = stocks.cell(rowIndex, sma8Col).value
                    if (first != None and second != None):
                        try:
                            numerator = abs(first - second)
                            stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                            twentifirst = numerator
                        except Exception as ex:
                            print(ex)

                elif (colIndex == diffall):
                    try:
                        numerator = (first + second + third + fourth + fifth + sixth + seventh + eighth + ninth + tenth + eleventh + twelveth  +thirteenth + fourtheenth + fifteenth + sixteenth + seventeenth + eigtheenth + nineteenth + twentieth + twentifirst)/21
                        stocks.cell(rowIndex, colIndex).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(numerator, rowIndex, numerator)
                        # first = 0
                        # second = 0
                        # third = 0
                        # fourth = 0
                        # fifth = 0
                        # sixth = 0

                        # seventh = 0
                        # eighth = 0
                        # ninth = 0
                        # tenth = 0
                        # eleventh = 0

                        # twelveth = 0
                        # thirteenth = 0
                        # fourtheenth = 0
                        # fifteenth = 0

                        # sixteenth = 0
                        # seventeenth = 0
                        # eigtheenth = 0

                        # nineteenth = 0
                        # twentieth = 0

                        # twentifirst = 0
                    except Exception as ex:
                        print(ex)

                elif (colIndex == volumeAvg200Col): 
                    rangeToUse = getNumDaysStr(date, 200, 0)
                    url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/chart/"+rangeToUse+"?chartCloseOnly=true&token="+token
                    print(url)
                    data = (requests.get(url).json())
                    data.reverse()
                    avg, stddev, volclose = calcAvgVol(data, date, 200)

                    if (avg != None):
                        stocks.cell(rowIndex, volumeAvg200Col).value = '=IF($BL$3="yes", MAX(0, ({0} - {1}) / {2}), {3})'.format(volclose, avg, stddev, avg)

                    avg, stddev, volclose = calcAvgVol(data, date, 100)

                    if (avg != None):
                        stocks.cell(rowIndex, volumeAvg100Col).value = '=IF($BL$3="yes", MAX(0, ({0} - {1}) / {2}), {3})'.format(volclose, avg, stddev, avg)

                    avg, stddev, volclose = calcAvgVol(data, date, 50)

                    if (avg != None):
                        stocks.cell(rowIndex, volumeAvg50Col).value = '=IF($BL$3="yes", MAX(0, ({0} - {1}) / {2}), {3})'.format(volclose, avg, stddev, avg)

                    avg, stddev, volclose = calcAvgVol(data, date, 20)

                    if (avg != None):
                        stocks.cell(rowIndex, volumeAvg20Col).value = '=IF($BL$3="yes", MAX(0, ({0} - {1}) / {2}), {3})'.format(volclose, avg, stddev, avg)


                    avg, stddev, volclose = calcAvgVol(data, date, 10)

                    if (avg != None):
                        stocks.cell(rowIndex, volumeAvg10Col).value = '=IF($BL$3="yes", MAX(0, ({0} - {1}) / {2}), {3})'.format(volclose, avg, stddev, avg)

                    avg, stddev, volclose = calcAvgVol(data, date, 5)

                    if (avg != None):
                        stocks.cell(rowIndex, volumeAvg5Col).value = '=IF($BL$3="yes", MAX(0, ({0} - {1}) / {2}), {3})'.format(volclose, avg, stddev, avg)

                elif (colIndex == rsi14Col):
                    try:
                        rangeToUse = getNumDaysStr(date, 0, 0)
                        url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/rsi?range="+rangeToUse+"&indicatorOnly=false&input1=14&token="+token
                        calc = calcIndicator(url, date, 0)
                        if (calc != None):
                            stocks.cell(rowIndex, colIndex).value = calc
                    except Exception as ex:
                        print(ex)


                elif (colIndex == accuDistLineCol):
                    try:
                        rangeToUse = getNumDaysStr(date, 14, 0)
                        url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/ad?range="+rangeToUse+"&indicatorOnly=false&token="+token
                        calc = calcIndicator(url, date, 0)
                        if (calc != None):
                            stocks.cell(rowIndex, colIndex).value = calc
                    except Exception as ex:
                        print(ex)

                elif (colIndex == accuDistOscCol):
                    try:
                        rangeToUse = getNumDaysStr(date, 5, 0)
                        url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/adosc?range="+rangeToUse+"&indicatorOnly=false&input1=2&input2=5&token="+token
                        calc = calcIndicator(url, date, 0)
                        if (calc != None):
                            stocks.cell(rowIndex, colIndex).value = calc
                    except Exception as ex:
                        print(ex)


                elif (colIndex == avgAtrCol):
                    try:
                        rangeToUse = getNumDaysStr(date, 14+50+14, 0)
                        url = "https://cloud.iexapis.com/stable/stock/" + symbol + "/indicator/atr?range="+rangeToUse+"&indicatorOnly=false&input1=14&token="+token
                        print(date, rangeToUse, url)
                        atrAvg, atr = calcAtr(url, date, 14, 50)
                        if (atr != None):
                            stocks.cell(rowIndex, avgAtrCol).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(atrAvg, rowIndex, atrAvg)
                            stocks.cell(rowIndex, atrCol).value = '=IF($Z$3="yes", {0}/$X{1}, {2})'.format(atr, rowIndex, atr)
                            # atrCol = column_index_from_string("BV")
                            # avgAtrCol = column_index_from_string("BW")
                            # deltaAvgAtr_AtrCol = column_index_from_string("BX")
                            avgc = get_column_letter(avgAtrCol)
                            atrc = get_column_letter(atrCol)
                            datrc = get_column_letter(deltaAvgAtr_AtrCol)
                            stocks.cell(rowIndex, deltaAvgAtr_AtrCol).value = '=${0}${1} - ${2}${3}'.format(avgc, rowIndex, atrc, rowIndex)
                    except Exception as ex:
                        print(ex)



                #linregslope
            colIndex += 1

        rowIndex += 1
        #xl.save(themesxl)

    try:
        xl.save(themesxl)
    except Exception as ex:
        failed = True
        while(failed):
            print(ex)
            str = input("Save failed, try again...? Press any key to try again")
            try:
                xl.save(themesxl)
                failed = False
            except Exception as ex:
                failed = True





def main():
    print("Script started")

if __name__ == "__main__":
    parser=argparse.ArgumentParser()

    parser.add_argument('--computeWL', '-wl', help='Compute Watchlist Data', type=str, default="y")
    parser.add_argument('--computeStocks', '-st', help='Compute highs', type=str, default="y")

    args=parser.parse_args()

    if (args.computeStocks == "y"):
        sheet = "Stocks"
        #sheet = "Watchlists"
        minr = 101
        maxr = 110
        #loadIndicators(sheet, minr, maxr)
        #def getStockData(sht, calcShOut, calcFloat, calcCountry, calcSector, calcIndustry, minrow, maxrow):
        calcShOut = True
        calcFloat = True
        calcCountry = True
        calcSector = True
        calcIndustry = True
        minr = 6
        maxr = 45
        getStockData(sheet, True, True, True, True, True, minr, maxr, False)

    if (args.computeWL == "y"):
        sheet = "Watchlists"
        #sheet = "Watchlists"
        minr = 6
        maxr = 160
        #loadIndicators(sheet, minr, maxr)
        #def getStockData(sht, calcShOut, calcFloat, calcCountry, calcSector, calcIndustry, minrow, maxrow):
        calcShOut = True
        calcFloat = True
        calcCountry = True
        calcSector = True
        calcIndustry = True
        minr = 6
        maxr = 160
        getStockData(sheet, True, True, True, True, True, minr, maxr, False)



    #/time-series/REPORTED_FINANCIALS/AAPL/10-Q?last=2

#pprint(stock.history(period="3mo"))



